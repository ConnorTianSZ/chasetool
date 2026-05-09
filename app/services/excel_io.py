"""Excel import / export service
Supports: WBS parsing, PGR mapping, quantity=0 skip, case-insensitive extensions
"""
from __future__ import annotations
import hashlib
import io
import json
import re
import sqlite3
from pathlib import Path
from typing import Any
import unicodedata

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import yaml

from app.db.connection import get_connection
from app.update_policy import bulk_update_fields, FACT_FIELDS
from app.services.material_view import clean_date_value, load_pgr_map

_MAPPING_PATH = Path(__file__).parent.parent.parent / "config" / "excel_column_mapping.yaml"
_DATE_FIELDS = {
    "order_date",
    "original_eta",
    "current_eta",
    "supplier_eta",
    "statical_delivery_date",
}


def _load_mapping() -> dict[str, list[str]]:
    with open(_MAPPING_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


_WBS_RE = re.compile(r"^([A-Za-z]\.\d+)\.(\d+)", re.IGNORECASE)


def parse_wbs(wbs: str) -> tuple[str | None, str | None]:
    if not wbs:
        return None, None
    m = _WBS_RE.match(wbs.strip())
    if m:
        return m.group(1), m.group(2)
    return None, None


def _normalize(text: str) -> str:
    text = str(text).strip()
    text = unicodedata.normalize("NFKC", text)
    return text.lower()


def _build_header_map(headers: list[str], mapping: dict[str, list[str]]) -> dict[str, str]:
    alias_index: dict[str, str] = {}
    for field, aliases in mapping.items():
        for alias in aliases:
            alias_index[_normalize(alias)] = field
    result: dict[str, str] = {}
    for h in headers:
        norm = _normalize(h)
        if norm in alias_index:
            result[h] = alias_index[norm]
    return result


def _file_hash(path: Path) -> str:
    h = hashlib.md5()
    h.update(path.read_bytes())
    return h.hexdigest()


def _is_excel(filename: str) -> bool:
    return filename.lower().endswith((".xlsx", ".xls"))


def import_excel(file_path: str | Path, project_id: str = "default") -> dict:
    path = Path(file_path)
    if not _is_excel(path.name):
        raise ValueError("Only .xlsx / .xls files are supported")

    mapping = _load_mapping()
    pgr_map = load_pgr_map()

    df = pd.read_excel(path, dtype=str, keep_default_na=False)
    df.columns = [str(c).strip() for c in df.columns]

    header_map = _build_header_map(list(df.columns), mapping)

    rows_added = 0
    rows_updated = 0
    rows_skipped = 0
    errors: list[dict] = []

    conn = get_connection(project_id)
    try:
        for idx, row in df.iterrows():
            raw = dict(row)

            po_col   = next((c for c, f in header_map.items() if f == "po_number"), None)
            item_col = next((c for c, f in header_map.items() if f == "item_no"),   None)

            po   = str(raw.get(po_col,   "")).strip() if po_col   else ""
            item = str(raw.get(item_col, "")).strip() if item_col else ""

            if not po or not item or po in ("nan", "") or item in ("nan", ""):
                errors.append({"row": int(idx) + 2, "reason": "missing PO or item number", "data": str(raw)})
                rows_skipped += 1
                continue

            qty_col = next((c for c, f in header_map.items() if f == "quantity"), None)
            if qty_col:
                qty_str = str(raw.get(qty_col, "")).strip()
                try:
                    if float(qty_str) == 0:
                        rows_skipped += 1
                        continue
                except (ValueError, TypeError):
                    pass

            data: dict[str, Any] = {}
            extra: dict[str, Any] = {}

            for col, val in raw.items():
                val_str = str(val).strip() if val != "" else None
                if val_str in ("nan", ""):
                    val_str = None

                if col in header_map:
                    field = header_map[col]
                    data[field] = val_str
                else:
                    if val_str:
                        extra[col] = val_str

            wbs = data.get("wbs_element")
            if wbs:
                project_no, station_no = parse_wbs(wbs)
                data["project_no"] = project_no
                data["station_no"] = station_no

            pg = (data.get("purchasing_group") or "").strip().upper()
            if pg and pg in pgr_map:
                pgr_info = pgr_map[pg]
                if not data.get("buyer_name"):
                    data["buyer_name"]  = pgr_info.get("name")
                if not data.get("buyer_email"):
                    data["buyer_email"] = pgr_info.get("email")

            for field in _DATE_FIELDS:
                if field in data:
                    data[field] = clean_date_value(data[field])

            if data.get("current_eta") and not data.get("original_eta"):
                data["original_eta"] = data["current_eta"]

            data["extra_json"] = json.dumps(extra, ensure_ascii=False) if extra else None

            cur = conn.execute(
                "SELECT id FROM materials WHERE po_number=? AND item_no=?", (po, item)
            )
            existing = cur.fetchone()

            if existing is None:
                cols = list(data.keys())
                placeholders = ", ".join(["?"] * len(cols))
                conn.execute(
                    "INSERT INTO materials (" + ", ".join(cols) + ") "
                    "VALUES (" + placeholders + ")",
                    [data[c] for c in cols],
                )
                rows_added += 1
            else:
                mat_id = existing[0]
                fact_updates = {
                    k: v for k, v in data.items()
                    if k in FACT_FIELDS and k not in ("po_number", "item_no")
                }
                semantic_updates = {
                    k: v for k, v in data.items()
                    if k not in FACT_FIELDS and k != "extra_json"
                }

                if fact_updates:
                    set_clause = ", ".join(k + "=?" for k in fact_updates)
                    conn.execute(
                        "UPDATE materials SET " + set_clause + " WHERE id=?",
                        [*fact_updates.values(), mat_id],
                    )

                if semantic_updates:
                    bulk_update_fields(conn, mat_id, semantic_updates, source="manual_import")

                if data.get("extra_json"):
                    conn.execute(
                        "UPDATE materials SET extra_json=? WHERE id=?",
                        (data["extra_json"], mat_id),
                    )

                rows_updated += 1

        fhash = _file_hash(path)
        from datetime import datetime
        conn.execute(
            "INSERT INTO imports "
            "(file_path, file_hash, rows_added, rows_updated, rows_skipped, errors_json, imported_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                str(path), fhash,
                rows_added, rows_updated, rows_skipped,
                json.dumps(errors, ensure_ascii=False),
                datetime.now().isoformat(sep=" ", timespec="seconds"),
            ),
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "rows_added":   rows_added,
        "rows_updated": rows_updated,
        "rows_skipped": rows_skipped,
        "errors":       errors,
    }


def export_back(
    source_path: Path,
    dest_path: Path | None = None,
    project_id: str = "default",
) -> Path:
    """Write current_eta and status back into a copy of the source Excel."""
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    mapping  = _load_mapping()
    hmap     = _build_header_map(headers, mapping)

    po_col_idx   = next((i + 1 for i, h in enumerate(headers) if hmap.get(h) == "po_number"), None)
    item_col_idx = next((i + 1 for i, h in enumerate(headers) if hmap.get(h) == "item_no"),   None)
    eta_col_idx  = next((i + 1 for i, h in enumerate(headers) if hmap.get(h) == "current_eta"), None)
    stat_col_idx = next((i + 1 for i, h in enumerate(headers) if hmap.get(h) == "status"),    None)

    if not po_col_idx or not item_col_idx:
        raise ValueError("Cannot find PO / Item columns in source file")

    conn = get_connection(project_id)
    try:
        for row_idx in range(2, ws.max_row + 1):
            po   = str(ws.cell(row_idx, po_col_idx).value or "").strip()
            item = str(ws.cell(row_idx, item_col_idx).value or "").strip()
            if not po or not item:
                continue
            db_row = conn.execute(
                "SELECT current_eta, status FROM materials WHERE po_number=? AND item_no=?",
                (po, item),
            ).fetchone()
            if db_row:
                if eta_col_idx:
                    ws.cell(row_idx, eta_col_idx).value = db_row["current_eta"]
                if stat_col_idx:
                    ws.cell(row_idx, stat_col_idx).value = db_row["status"]
    finally:
        conn.close()

    out = dest_path or source_path.with_stem(source_path.stem + "_updated")
    wb.save(out)
    return out


# ── 完整数据库导出列定义 ─────────────────────────────────────────────────────
# (Excel列头, 数据库字段名)
_FULL_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Plant",                                      "plant"),
    ("Supplier",                                   "supplier_code"),
    ("Name 1",                                     "supplier"),
    ("WBS Element",                                "wbs_element"),
    ("Purchasing Group",                           "purchasing_group"),
    ("Document number",                            "po_number"),
    ("Item",                                       "item_no"),
    ("Document Date",                              "order_date"),
    ("Statical Delivery Date",                     "statical_delivery_date"),
    ("First confirmed DelDate",                    "original_eta"),
    ("Confirmed Del. Date",                        "current_eta"),
    ("Material",                                   "part_no"),
    ("Short Text",                                 "description"),
    ("Manufacturer",                               "manufacturer"),
    ("Manufacturer Part No.",                      "manufacturer_part_no"),
    ("Order Quantity",                             "quantity"),
    ("Open quantity GR",                           "open_quantity_gr"),
    ("Order Unit",                                 "unit"),
    ("Net Order Price",                            "net_order_price"),
    ("Currency for Purchasing document in docu",   "currency"),
    ("Net Order Value",                            "net_order_value"),
    ("Position Text1",                             "position_text1"),
    ("Position Text2",                             "position_text2"),
    # ── 催货附加字段 ──
    ("供应商回复交期",                              "supplier_eta"),
    ("催货状态",                                    "status"),
    ("催货次数",                                    "chase_count"),
    ("最后催货时间",                                "last_chased_at"),
    ("供应商回复备注",                              "supplier_remarks"),
]

# 重点行样式
_FOCUS_FILL = PatternFill(fill_type="solid", fgColor="FFF3C6")   # 浅黄底

# 表头样式
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")  # 淡蓝底
_HEADER_FONT = Font(bold=True)


def _apply_focus_row(ws, row_idx: int, max_col: int) -> None:
    """将指定行设置为重点行样式（加粗 + 背景色）。"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row_idx, col)
        cell.font = Font(bold=True, name=(cell.font.name or "Calibri"))
        cell.fill = _FOCUS_FILL


def export_full_db(project_id: str = "default") -> bytes:
    """将当前项目数据库的全部物料导出为 Excel 字节流。

    重点物料（is_focus=1）整行加粗并填充黄色背景。
    返回值为 .xlsx 文件的原始字节，可直接作为 HTTP 响应体。
    """
    conn = get_connection(project_id)
    try:
        rows = conn.execute(
            "SELECT * FROM materials ORDER BY "
            "CASE WHEN (open_quantity_gr IS NOT NULL AND CAST(open_quantity_gr AS REAL)=0) THEN 1 ELSE 0 END ASC, "
            "date(current_eta) ASC, po_number ASC, item_no ASC"
        ).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materials"

    headers   = [col_name for col_name, _ in _FULL_EXPORT_COLUMNS]
    db_fields = [field    for _, field    in _FULL_EXPORT_COLUMNS]

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # 写数据
    for row_idx, row in enumerate(rows, 2):
        r = dict(row)
        is_focus = bool(r.get("is_focus"))
        for col_idx, field in enumerate(db_fields, 1):
            ws.cell(row_idx, col_idx, r.get(field))
        if is_focus:
            _apply_focus_row(ws, row_idx, len(headers))

    # 列宽自适应（简单估算）
    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 2, 10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 追加催货字段列定义（用于 export_chase_append）────────────────────────────
_CHASE_APPEND_COLUMNS: list[tuple[str, str]] = [
    ("供应商回复交期",  "supplier_eta"),
    ("催货状态",        "status"),
    ("催货次数",        "chase_count"),
    ("最后催货时间",    "last_chased_at"),
    ("供应商回复备注",  "supplier_remarks"),
]


def export_chase_append(
    source_path: str | Path,
    project_id: str = "default",
) -> Path:
    """在原始 Excel 末尾追加催货信息列，保存为 <原名>-chase.xlsx。

    - 不改变原始文件格式（原有列保持原样）。
    - 以 PO号 + Item号 为联合键匹配数据库记录。
    - 重点物料（is_focus=1）整行加粗并填充黄色背景。
    - 输出文件与源文件同目录，命名为 <stem>-chase.xlsx。
    """
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(f"源文件不存在：{source_path}")

    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    # ── 找 PO / Item 列索引（基于列头名匹配映射表）──
    src_headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    mapping = _load_mapping()
    hmap = _build_header_map(src_headers, mapping)

    po_col_idx   = next((i + 1 for i, h in enumerate(src_headers) if hmap.get(h) == "po_number"), None)
    item_col_idx = next((i + 1 for i, h in enumerate(src_headers) if hmap.get(h) == "item_no"),   None)

    if not po_col_idx or not item_col_idx:
        raise ValueError("源文件中无法识别 PO号 或 行号 列，请检查列头。")

    # ── 原表最后一列之后追加新列头 ──
    orig_max_col    = ws.max_column
    chase_start_col = orig_max_col + 1

    chase_headers = [col_name for col_name, _ in _CHASE_APPEND_COLUMNS]
    chase_fields  = [field    for _, field    in _CHASE_APPEND_COLUMNS]

    for offset, header in enumerate(chase_headers):
        col_idx = chase_start_col + offset
        cell = ws.cell(1, col_idx, header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # ── 从数据库批量读取所有物料的催货数据 ──
    conn = get_connection(project_id)
    try:
        db_rows = conn.execute(
            "SELECT po_number, item_no, supplier_eta, status, chase_count, "
            "last_chased_at, supplier_remarks, is_focus FROM materials"
        ).fetchall()
    finally:
        conn.close()

    # 建立 (po, item) → row dict 的快速查找表
    db_index: dict[tuple[str, str], dict] = {}
    for r in db_rows:
        d = dict(r)
        key = (str(d["po_number"] or "").strip(), str(d["item_no"] or "").strip())
        db_index[key] = d

    # ── 逐行填充催货数据 ──
    total_col = chase_start_col + len(chase_headers) - 1
    for row_idx in range(2, ws.max_row + 1):
        po   = str(ws.cell(row_idx, po_col_idx).value or "").strip()
        item = str(ws.cell(row_idx, item_col_idx).value or "").strip()
        if not po and not item:
            continue

        db_row = db_index.get((po, item))
        if db_row:
            for offset, field in enumerate(chase_fields):
                ws.cell(row_idx, chase_start_col + offset, db_row.get(field))
            if db_row.get("is_focus"):
                _apply_focus_row(ws, row_idx, total_col)

    # 新增列宽
    for offset, header in enumerate(chase_headers):
        col_letter = openpyxl.utils.get_column_letter(chase_start_col + offset)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 14)

    out_path = source_path.with_name(source_path.stem + "-chase.xlsx")
    wb.save(out_path)
    return out_path
